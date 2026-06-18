from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font


def export_contracts_excel(contracts) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Contracten"

    headers = [
        "Contract Nr", "Customer", "Kostenplaats", "Description",
        "Start Date", "End Date", "Value", "Status", "Notes",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, contract in enumerate(contracts, 2):
        ws.cell(row=row_num, column=1, value=contract.contract_number)
        ws.cell(row=row_num, column=2, value=contract.customer_name)
        ws.cell(row=row_num, column=3, value=contract.kostenplaats)
        ws.cell(row=row_num, column=4, value=contract.description)
        ws.cell(row=row_num, column=5, value=contract.start_date.isoformat() if contract.start_date else None)
        ws.cell(row=row_num, column=6, value=contract.end_date.isoformat() if contract.end_date else None)
        ws.cell(row=row_num, column=7, value=contract.contract_value)
        ws.cell(row=row_num, column=8, value=contract.status)
        ws.cell(row=row_num, column=9, value=contract.notes)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
